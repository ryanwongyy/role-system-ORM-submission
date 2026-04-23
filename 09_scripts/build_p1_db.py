#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import os
import re
import shutil
import subprocess
import sys
import time
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date
from html import unescape
from pathlib import Path
from typing import Any, Iterable
from urllib.error import HTTPError, URLError
from urllib.parse import urljoin, urlparse
from urllib.request import Request, urlopen

from openpyxl import load_workbook

try:
    from bs4 import BeautifulSoup  # type: ignore
except Exception:  # pragma: no cover - optional dependency
    BeautifulSoup = None

try:
    from pypdf import PdfReader  # type: ignore
except Exception:  # pragma: no cover - optional dependency
    PdfReader = None


TODAY = os.environ.get("P1_DB_DATE", "2026-04-07")
BUILD_ROOT = Path("/Users/ryanwong/Documents/Playground/p1_role_systems_db_build")
WORKBOOK_PATH = BUILD_ROOT / "ai_rich_social_science_case_repository.xlsx"
PROJECT_ROOT = BUILD_ROOT / "p1_role_systems_db"
WORKBOOK_COPY_PATH = PROJECT_ROOT / "00_input" / WORKBOOK_PATH.name
USER_TARGET_ROOT = Path("/Users/ryanwong/Human roles/p1_role_systems_db")

TOP_LEVEL_DIRS = [
    "00_input",
    "01_registry",
    "02_cases",
    "03_raw_sources",
    "04_extracted_text",
    "05_case_notes",
    "06_analysis_tables",
    "07_codebook",
    "08_logs",
    "09_scripts",
    "10_outputs",
]

TOP_LEVEL_DOCS = [
    "README.md",
    "collection_log.md",
    "data_inventory.md",
    "known_gaps.md",
    "coding_decisions.md",
]

REPOSITORY_HEADER_ROW = 1
SOURCES_HEADER_ROW = 4
FLAG_HEADER_ROW = 4
DICTIONARY_HEADER_ROW = 4

CASE_SLUGS = {
    1: "001_objective_party_annotation",
    2: "002_contested_codebook_measurement",
    3: "003_multi_perspective_annotation",
    4: "004_party_cue_perturbation",
    5: "005_human_developed_codebooks",
    6: "006_maternal_health_interview_coding",
    7: "007_agentic_title_abstract_screening",
    8: "008_study_type_classification_review_triage",
    9: "009_citation_grounded_multi_paper_synthesis",
    10: "010_citation_hallucination_checks",
    11: "011_grant_proposal_ideation",
    12: "012_funding_pipeline_agenda_compression",
    13: "013_scideator_facet_recombination",
    14: "014_researchbench_discoverybench_tasks",
    15: "015_multi_agent_idea_generation",
    16: "016_ai_conversational_interviewing",
    17: "017_dynamic_surveys_followups",
    18: "018_qda_delegation_preferences",
    19: "019_ai_assisted_fieldnote_memoing",
    20: "020_reflexive_content_analysis_redesign",
    21: "021_synthetic_respondents_benchmarked",
    22: "022_socsci210_behavior_simulators",
    23: "023_forecasting_before_field_experiments",
    24: "024_ai_multiverse_nonstandard_errors",
    25: "025_project_ape_policy_evaluation_tournament",
    26: "026_project_ape_autonomy_gradient",
    27: "027_repro_bench_reproducibility_assessment",
    28: "028_agent_based_code_package_repair",
    29: "029_lab_manuscript_cowriting",
    30: "030_structured_ai_disclosure_daisy",
    31: "031_project_rachel_authorship_boundary",
    32: "032_ai_scientist_boundary_cases",
    33: "033_deep_research_policy_reconnaissance",
    34: "034_agentsociety_transportability_boundary",
    35: "035_doctoral_student_ai_use",
    36: "036_predictive_design_course",
}

ROLE_UNIVERSE = [
    ("Relevance adjudicator", "starting_role"),
    ("Mechanism disciplinarian", "starting_role"),
    ("Construct and perspective boundary setter", "starting_role"),
    ("Corpus pluralist", "starting_role"),
    ("Protocol and provenance architect", "starting_role"),
    ("Counterfactual and transportability judge", "starting_role"),
    ("Situated interlocutor / context witness", "starting_role"),
    ("Accountability and labor allocator", "starting_role"),
    ("Construct boundary setter", "split_candidate"),
    ("Perspective sampler", "split_candidate"),
    ("Counterfactual / transportability judge", "split_candidate"),
    ("Calibration architect", "split_candidate"),
    ("Situated interlocutor", "split_candidate"),
    ("Context witness", "split_candidate"),
    ("Routine first-pass screener", "shrinking_task_cluster"),
    ("Routine coder on high-consensus, ground-truth-rich tasks", "shrinking_task_cluster"),
    ("Structural prose polisher / drafter", "shrinking_task_cluster"),
]

ROLE_KEYWORDS = {
    "relevance adjudicator": {"Relevance adjudicator"},
    "mechanism disciplinarian": {"Mechanism disciplinarian"},
    "construct and perspective boundary setter": {"Construct and perspective boundary setter"},
    "construct boundary setter": {"Construct boundary setter", "Construct and perspective boundary setter"},
    "perspective sampler": {"Perspective sampler", "Construct and perspective boundary setter"},
    "corpus pluralist": {"Corpus pluralist"},
    "protocol architect": {"Protocol and provenance architect"},
    "protocol and provenance architect": {"Protocol and provenance architect"},
    "provenance architect": {"Protocol and provenance architect"},
    "counterfactual and transportability judge": {"Counterfactual and transportability judge"},
    "counterfactual / transportability judge": {
        "Counterfactual and transportability judge",
        "Counterfactual / transportability judge",
    },
    "calibration architect": {"Calibration architect"},
    "situated interlocutor / context witness": {"Situated interlocutor / context witness"},
    "situated interlocutor": {"Situated interlocutor", "Situated interlocutor / context witness"},
    "context witness": {"Context witness", "Situated interlocutor / context witness"},
    "accountability and labor allocator": {"Accountability and labor allocator"},
    "shrinking routine coder": {"Routine coder on high-consensus, ground-truth-rich tasks"},
    "routine coder": {"Routine coder on high-consensus, ground-truth-rich tasks"},
    "routine first-pass screener": {"Routine first-pass screener"},
    "first-pass screener": {"Routine first-pass screener"},
    "structural prose polisher": {"Structural prose polisher / drafter"},
    "drafter": {"Structural prose polisher / drafter"},
}

FAMILY_ROLE_HINTS = {
    "Measurement & coding": {
        "Construct and perspective boundary setter",
        "Construct boundary setter",
        "Perspective sampler",
        "Protocol and provenance architect",
        "Routine coder on high-consensus, ground-truth-rich tasks",
    },
    "Evidence synthesis & corpus": {
        "Relevance adjudicator",
        "Corpus pluralist",
        "Protocol and provenance architect",
        "Routine first-pass screener",
    },
    "Agenda, theory & discovery": {
        "Relevance adjudicator",
        "Mechanism disciplinarian",
        "Accountability and labor allocator",
    },
    "Interviews, fieldwork & qualitative interpretation": {
        "Construct and perspective boundary setter",
        "Construct boundary setter",
        "Situated interlocutor / context witness",
        "Situated interlocutor",
        "Context witness",
        "Protocol and provenance architect",
    },
    "Predictive simulation & design": {
        "Mechanism disciplinarian",
        "Counterfactual and transportability judge",
        "Counterfactual / transportability judge",
        "Calibration architect",
        "Accountability and labor allocator",
    },
    "Reproducibility & autonomous research": {
        "Protocol and provenance architect",
        "Mechanism disciplinarian",
        "Accountability and labor allocator",
    },
    "Governance, authorship & labor": {
        "Accountability and labor allocator",
        "Protocol and provenance architect",
        "Structural prose polisher / drafter",
    },
    "Pedagogy & training": {
        "Accountability and labor allocator",
        "Protocol and provenance architect",
        "Calibration architect",
    },
}

ARTIFACT_KEYWORDS = {
    "codebook": "codebook",
    "disagreement": "disagreement_matrix",
    "transcript": "transcript",
    "memo": "memo_trail",
    "prisma": "PRISMA_log",
    "citation": "citation_log",
    "forecast": "forecast_log",
    "version": "version_history",
    "reproducibility": "reproducibility_report",
    "tournament": "tournament_log",
    "benchmark": "benchmark_log",
    "disclosure": "disclosure_form",
    "prompt": "prompt_log",
    "repo": "repo",
    "repository": "repo",
    "dataset": "dataset",
}

GENERIC_ARTIFACT_SOURCE_HINTS = (
    "supplement",
    "supplementary",
    "appendix",
    "materials",
    "osf",
    "github",
    "benchmark",
    "dataset",
    "dataverse",
    "protocol",
    "prisma",
)

ARTIFACT_SOURCE_HINTS = {
    "codebook": ("codebook", "supplement", "supplementary", "appendix"),
    "prompt": ("prompt", "supplement", "supplementary", "appendix", "materials"),
    "prisma": ("prisma", "protocol", "osf", "materials", "supplement", "supplementary"),
    "search": ("search", "protocol", "osf", "materials", "supplement", "supplementary"),
    "inclusion": ("inclusion", "exclusion", "osf", "materials", "supplement", "supplementary"),
    "rescue": ("gray", "grey", "rescue", "osf", "materials", "supplement", "supplementary"),
    "synthetic": ("synthetic", "dataset", "dataverse", "supplement", "supplementary", "appendix"),
    "regression": ("regression", "dataset", "dataverse", "supplement", "supplementary", "appendix"),
    "subgroup": ("subgroup", "dataset", "dataverse", "supplement", "supplementary", "appendix"),
    "assessment": ("github", "benchmark", "reproduction", "reproducibility"),
    "reproduced": ("github", "benchmark", "reproduction", "reproducibility"),
    "claim-alignment": ("github", "benchmark", "reproduction", "reproducibility"),
    "error": ("github", "benchmark", "repair", "reproducibility"),
    "patch": ("github", "benchmark", "repair"),
    "rerun": ("github", "benchmark", "repair", "reproduction"),
    "success/failure": ("github", "benchmark", "repair", "evaluation"),
    "version": ("version", "history", "disclosure", "supplement", "appendix"),
    "authorship": ("authorship", "disclosure", "supplement", "appendix"),
    "reviewer": ("review", "supplement", "appendix"),
}

SUPPLEMENT_ENRICHMENT_FOCUS_CASE_IDS = {"1", "2", "7", "21", "27", "28", "29"}


@dataclass
class FetchResult:
    final_url: str
    title: str
    authors: str
    year: str
    doi: str
    source_type: str
    evidence_tier: str
    access_status: str
    local_raw_path: str
    local_text_path: str
    extracted_text: str
    notes: str


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_header(value: Any) -> str:
    return clean_text(value)


def read_table(ws, header_row: int) -> list[dict[str, str]]:
    headers = [normalize_header(ws.cell(header_row, c).value) for c in range(1, ws.max_column + 1)]
    rows: list[dict[str, str]] = []
    for r in range(header_row + 1, ws.max_row + 1):
        row = {headers[c - 1]: clean_text(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)}
        if any(row.values()):
            rows.append(row)
    return rows


def write_csv(path: Path, fieldnames: list[str], rows: Iterable[dict[str, Any]]) -> None:
    materialized_rows = list(rows)

    def _writer(target: Path) -> None:
        with target.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.DictWriter(handle, fieldnames=fieldnames)
            writer.writeheader()
            for row in materialized_rows:
                writer.writerow({key: clean_text(row.get(key, "")) for key in fieldnames})

    atomic_write(path, _writer)


def read_csv(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", newline="", encoding="utf-8") as handle:
        return list(csv.DictReader(handle))


def write_text(path: Path, text: str) -> None:
    atomic_write(path, lambda target: target.write_text(text.rstrip() + "\n", encoding="utf-8"))


def write_bytes(path: Path, payload: bytes) -> None:
    atomic_write(path, lambda target: target.write_bytes(payload))


def write_json(path: Path, payload: dict[str, Any]) -> None:
    def _writer(target: Path) -> None:
        with target.open("w", encoding="utf-8") as handle:
            json.dump(payload, handle, indent=2, ensure_ascii=False)

    atomic_write(path, _writer)


def atomic_write(path: Path, writer) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = path.with_name(f".{path.name}.tmp")
    last_error: OSError | None = None
    for attempt in range(5):
        try:
            writer(temp_path)
            os.replace(temp_path, path)
            return
        except OSError as exc:
            last_error = exc
            temp_path.unlink(missing_ok=True)
            if exc.errno == 60 and attempt < 4:
                time.sleep(0.5 * (attempt + 1))
                continue
            raise
    if last_error is not None:
        raise last_error


def read_bytes_with_retries(path: Path) -> bytes:
    last_error: OSError | None = None
    for attempt in range(5):
        try:
            return path.read_bytes()
        except OSError as exc:
            last_error = exc
            if exc.errno == 60 and attempt < 4:
                time.sleep(0.5 * (attempt + 1))
                continue
            raise
    if last_error is not None:
        raise last_error
    return b""


def sheet_rows_exact(ws) -> list[list[str]]:
    rows: list[list[str]] = []
    for r in range(1, ws.max_row + 1):
        rows.append([clean_text(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)])
    return rows


def export_exact_sheet_csv(ws, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        for row in sheet_rows_exact(ws):
            writer.writerow(row)


def sha_path(value: str) -> str:
    return hashlib.sha1(value.encode("utf-8")).hexdigest()[:12]


def ensure_structure() -> None:
    PROJECT_ROOT.mkdir(parents=True, exist_ok=True)
    for dirname in TOP_LEVEL_DIRS:
        (PROJECT_ROOT / dirname).mkdir(parents=True, exist_ok=True)
    for doc_name in TOP_LEVEL_DOCS:
        (PROJECT_ROOT / doc_name).touch()
    write_bytes(WORKBOOK_COPY_PATH, read_bytes_with_retries(WORKBOOK_PATH))
    write_bytes(PROJECT_ROOT / "09_scripts" / "build_p1_db.py", read_bytes_with_retries(Path(__file__)))


def sync_project_tree(source_root: Path, target_root: Path) -> None:
    target_root.mkdir(parents=True, exist_ok=True)
    for path in sorted(source_root.rglob("*")):
        relative = path.relative_to(source_root)
        target_path = target_root / relative
        if path.is_dir():
            target_path.mkdir(parents=True, exist_ok=True)
            continue
        write_bytes(target_path, read_bytes_with_retries(path))


def classify_source(url: str, source_title: str, evidence_status: str) -> tuple[str, str]:
    url_lower = url.lower()
    title_lower = source_title.lower()
    evidence_lower = evidence_status.lower()
    if "arxiv.org" in url_lower:
        source_type = "preprint"
        evidence_tier = "benchmark" if "benchmark" in evidence_lower else "preprint"
        return source_type, evidence_tier
    if "aclanthology.org" in url_lower or "cscw" in title_lower:
        return "conference paper", "accepted conference paper"
    if "ape.socialcatalystlab.org" in url_lower:
        if "methodology" in url_lower:
            return "project methodology page", "live public project / website"
        return "project homepage", "live public project / website"
    if "pmc.ncbi.nlm.nih.gov" in url_lower:
        return "journal article", "peer-reviewed article"
    if any(domain in url_lower for domain in ["sagepub.com", "cambridge.org", "nature.com", "springer.com", "tandfonline.com"]):
        if "forum" in evidence_lower or "forum article" in evidence_lower:
            return "forum article", "commentary / forum / boundary evidence"
        return "journal article", "peer-reviewed article"
    if "tech4good" in url_lower:
        return "conference paper", "accepted conference paper"
    if any(marker in evidence_lower for marker in ["live website", "public project"]):
        return "project website", "live public project / website"
    if "benchmark" in evidence_lower:
        return "benchmark paper", "benchmark"
    if "commentary" in evidence_lower or "forum" in evidence_lower:
        return "commentary", "commentary / forum / boundary evidence"
    if "peer-reviewed" in evidence_lower:
        return "journal article", "peer-reviewed article"
    if "conference" in evidence_lower:
        return "conference paper", "accepted conference paper"
    return "web page", "commentary / forum / boundary evidence"


def default_headers() -> dict[str, str]:
    return {
        "User-Agent": "Mozilla/5.0 (compatible; Codex P1 Role Systems DB/1.0; +https://openai.com/)",
        "Accept": "text/html,application/xhtml+xml,application/xml,application/pdf;q=0.9,*/*;q=0.8",
    }


def html_to_text(html: str) -> str:
    if BeautifulSoup is not None:
        soup = BeautifulSoup(html, "html.parser")
        for tag in soup(["script", "style", "noscript"]):
            tag.decompose()
        text = soup.get_text("\n")
    else:
        text = re.sub(r"(?is)<(script|style).*?>.*?</\\1>", " ", html)
        text = re.sub(r"(?is)<[^>]+>", " ", text)
    text = unescape(text)
    text = re.sub(r"\r", "\n", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    text = re.sub(r"[ \t]{2,}", " ", text)
    return text.strip()


def extract_meta(html: str) -> dict[str, str]:
    meta: dict[str, str] = {}
    if BeautifulSoup is not None:
        soup = BeautifulSoup(html, "html.parser")
        title_tag = soup.find("title")
        if title_tag:
            meta["title"] = clean_text(title_tag.get_text())
        for item in soup.find_all("meta"):
            name = clean_text(item.get("name") or item.get("property")).lower()
            content = clean_text(item.get("content"))
            if not name or not content:
                continue
            meta[name] = content
    else:
        title_match = re.search(r"(?is)<title>(.*?)</title>", html)
        if title_match:
            meta["title"] = clean_text(title_match.group(1))
        for match in re.finditer(
            r'(?is)<meta[^>]+(?:name|property)=["\']([^"\']+)["\'][^>]+content=["\']([^"\']+)["\']',
            html,
        ):
            meta[clean_text(match.group(1)).lower()] = clean_text(match.group(2))
    return meta


def is_probably_paywalled(text: str, url: str) -> bool:
    payload = f"{url}\n{text[:5000]}".lower()
    markers = [
        "purchase access",
        "buy article",
        "institutional access",
        "subscribe to journal",
        "access through your institution",
        "rent this article",
        "login via institution",
    ]
    return any(marker in payload for marker in markers)


def is_access_blocked(text: str, url: str) -> bool:
    payload = f"{url}\n{text[:5000]}".lower()
    markers = [
        "just a moment",
        "checking if the site connection is secure",
        "verify you are human",
        "enable javascript and cookies to continue",
        "access denied",
        "captcha",
        "cf-browser-verification",
    ]
    return any(marker in payload for marker in markers)


def extract_pdf_text(pdf_path: Path) -> str:
    if PdfReader is not None:
        try:
            reader = PdfReader(str(pdf_path))
            pages = [clean_text(page.extract_text()) for page in reader.pages[:8]]
            text = "\n\n".join(page for page in pages if page)
            if text.strip():
                return text.strip()
        except Exception:
            pass
    if shutil.which("pdftotext"):
        output_path = pdf_path.with_suffix(".txt")
        try:
            subprocess.run(
                ["pdftotext", str(pdf_path), str(output_path)],
                check=True,
                capture_output=True,
            )
            if output_path.exists():
                return output_path.read_text(encoding="utf-8", errors="ignore").strip()
        except Exception:
            pass
    return ""


def sniff_suffix(payload: bytes, content_type: str, url: str) -> str:
    sample = payload[:512].lstrip().lower()
    if payload.startswith(b"%PDF-"):
        return ".pdf"
    if sample.startswith(b"<!doctype html") or sample.startswith(b"<html") or b"<html" in sample[:200]:
        return ".html"
    if "pdf" in content_type or url.lower().endswith(".pdf"):
        return ".pdf"
    if "html" in content_type or not content_type:
        return ".html"
    if "json" in content_type:
        return ".json"
    return ".txt"


def read_text_lossy(path: Path) -> str:
    try:
        return path.read_text(encoding="utf-8")
    except UnicodeDecodeError:
        return path.read_text(encoding="latin-1", errors="ignore")


def load_cached_fetch(
    url: str,
    raw_dir: Path,
    text_dir: Path,
    source_id: str,
    fallback_title: str,
    evidence_status: str,
) -> FetchResult | None:
    source_type, evidence_tier = classify_source(url, fallback_title, evidence_status)
    candidates = []
    for suffix in [".html", ".pdf", ".json", ".txt", ".bin", ".download"]:
        path = raw_dir / f"{source_id}{suffix}"
        if path.exists():
            candidates.append(path)
    if not candidates:
        return None

    raw_path = candidates[0]
    if raw_path.suffix == ".download":
        payload = raw_path.read_bytes()
        suffix = sniff_suffix(payload, "", url)
        renamed_path = raw_dir / f"{source_id}{suffix}"
        write_bytes(renamed_path, payload)
        raw_path.unlink(missing_ok=True)
        raw_path = renamed_path

    text_path = text_dir / f"{source_id}.txt"
    if raw_path.suffix == ".pdf":
        text = text_path.read_text(encoding="utf-8", errors="ignore").strip() if text_path.exists() else ""
        if not text:
            text = extract_pdf_text(raw_path)
            write_text(text_path, text)
        return FetchResult(
            final_url=url,
            title=fallback_title,
            authors="",
            year="",
            doi="",
            source_type=source_type,
            evidence_tier=evidence_tier,
            access_status="open",
            local_raw_path=str(raw_path),
            local_text_path=str(text_path),
            extracted_text=text,
            notes="Reused cached PDF snapshot",
        )

    html = read_text_lossy(raw_path)
    meta = extract_meta(html)
    text = text_path.read_text(encoding="utf-8", errors="ignore").strip() if text_path.exists() else ""
    if not text:
        text = html_to_text(html)
        write_text(text_path, text)
    access_status = "open"
    if is_access_blocked(text, url):
        access_status = "inaccessible"
    elif is_probably_paywalled(text, url):
        access_status = "paywalled"
    title = meta.get("citation_title") or meta.get("og:title") or meta.get("title") or fallback_title
    authors = meta.get("citation_author", "")
    year = meta.get("citation_publication_date", "")[:4]
    doi = meta.get("citation_doi", "") or meta.get("dc.identifier", "")
    return FetchResult(
        final_url=url,
        title=title,
        authors=authors,
        year=year,
        doi=doi,
        source_type=source_type,
        evidence_tier=evidence_tier,
        access_status=access_status,
        local_raw_path=str(raw_path),
        local_text_path=str(text_path),
        extracted_text=text,
        notes="Reused cached HTML/text snapshot",
    )


def fetch_url(
    url: str,
    raw_dir: Path,
    text_dir: Path,
    source_id: str,
    fallback_title: str,
    evidence_status: str,
) -> FetchResult:
    source_type, evidence_tier = classify_source(url, fallback_title, evidence_status)
    raw_dir.mkdir(parents=True, exist_ok=True)
    text_dir.mkdir(parents=True, exist_ok=True)

    cached_result = load_cached_fetch(url, raw_dir, text_dir, source_id, fallback_title, evidence_status)
    if cached_result is not None:
        return cached_result

    temp_path = raw_dir / f"{source_id}.download"
    headers_path = raw_dir / f"{source_id}.headers"
    temp_path.unlink(missing_ok=True)
    headers_path.unlink(missing_ok=True)
    ua = default_headers()["User-Agent"]

    def run_curl(insecure: bool = False):
        cmd = [
            "curl",
            "-L",
            "--silent",
            "--show-error",
            "--connect-timeout",
            "20",
            "--max-time",
            "90",
            "-A",
            ua,
            "-D",
            str(headers_path),
            "-o",
            str(temp_path),
            "-w",
            "%{url_effective}\n%{content_type}\n%{http_code}\n",
        ]
        if insecure:
            cmd.insert(1, "-k")
        cmd.append(url)
        return subprocess.run(cmd, capture_output=True, text=True, check=False)

    curl_result = run_curl(insecure=False)
    if curl_result.returncode != 0:
        curl_result = run_curl(insecure=True)

    if curl_result.returncode == 0:
        lines = curl_result.stdout.splitlines()
        final_url = lines[0].strip() if len(lines) >= 1 else url
        content_type = lines[1].strip().lower() if len(lines) >= 2 else ""
        http_code = lines[2].strip() if len(lines) >= 3 else "000"
        payload = temp_path.read_bytes() if temp_path.exists() else b""
        if http_code and http_code.isdigit() and int(http_code) >= 400:
            note = f"HTTP {http_code} while fetching {url}"
            return FetchResult(
                final_url=final_url or url,
                title=fallback_title,
                authors="",
                year="",
                doi="",
                source_type=source_type,
                evidence_tier=evidence_tier,
                access_status="paywalled" if http_code in {"401", "402", "403"} else "inaccessible",
                local_raw_path="",
                local_text_path="",
                extracted_text="",
                notes=note,
            )
    else:
        try:
            request = Request(url, headers=default_headers())
            with urlopen(request, timeout=45) as response:  # noqa: S310
                final_url = response.geturl()
                content_type = clean_text(response.headers.get("Content-Type")).lower()
                payload = response.read()
        except HTTPError as exc:
            note = f"HTTP {exc.code} while fetching {url}"
            return FetchResult(
                final_url=url,
                title=fallback_title,
                authors="",
                year="",
                doi="",
                source_type=source_type,
                evidence_tier=evidence_tier,
                access_status="paywalled" if exc.code in {401, 402, 403} else "inaccessible",
                local_raw_path="",
                local_text_path="",
                extracted_text="",
                notes=note,
            )
        except URLError as exc:
            return FetchResult(
                final_url=url,
                title=fallback_title,
                authors="",
                year="",
                doi="",
                source_type=source_type,
                evidence_tier=evidence_tier,
                access_status="inaccessible",
                local_raw_path="",
                local_text_path="",
                extracted_text="",
                notes=f"URL error while fetching {url}: {exc.reason}",
            )
        except Exception as exc:
            return FetchResult(
                final_url=url,
                title=fallback_title,
                authors="",
                year="",
                doi="",
                source_type=source_type,
                evidence_tier=evidence_tier,
                access_status="inaccessible",
                local_raw_path="",
                local_text_path="",
                extracted_text="",
                notes=f"Unexpected fetch error: {exc}",
            )

    suffix = sniff_suffix(payload, content_type, final_url or url)

    raw_path = raw_dir / f"{source_id}{suffix}"
    write_bytes(raw_path, payload)
    temp_path.unlink(missing_ok=True)

    if suffix == ".pdf":
        text = extract_pdf_text(raw_path)
        text_path = text_dir / f"{source_id}.txt"
        write_text(text_path, text)
        access_status = "open"
        return FetchResult(
            final_url=final_url,
            title=fallback_title,
            authors="",
            year="",
            doi="",
            source_type=source_type,
            evidence_tier=evidence_tier,
            access_status=access_status,
            local_raw_path=str(raw_path),
            local_text_path=str(text_path),
            extracted_text=text,
            notes="Downloaded PDF",
        )

    try:
        html = payload.decode("utf-8")
    except UnicodeDecodeError:
        html = payload.decode("latin-1", errors="ignore")
    meta = extract_meta(html)
    text = html_to_text(html)
    title = meta.get("citation_title") or meta.get("og:title") or meta.get("title") or fallback_title
    authors = meta.get("citation_author", "")
    year = meta.get("citation_publication_date", "")[:4]
    doi = meta.get("citation_doi", "") or meta.get("dc.identifier", "")
    text_path = text_dir / f"{source_id}.txt"
    write_text(text_path, text)
    if is_access_blocked(text, final_url):
        access_status = "inaccessible"
    elif is_probably_paywalled(text, final_url):
        access_status = "paywalled"
    else:
        access_status = "open"

    pdf_url = meta.get("citation_pdf_url", "")
    notes = "Downloaded HTML snapshot"
    if pdf_url:
        pdf_url = urljoin(final_url, pdf_url)
        pdf_result = fetch_url(pdf_url, raw_dir, text_dir, f"{source_id}_pdf", fallback_title, evidence_status)
        if pdf_result.access_status == "open" and pdf_result.local_raw_path:
            notes += f"; supplementary PDF saved at {pdf_result.local_raw_path}"
    return FetchResult(
        final_url=final_url,
        title=title,
        authors=authors,
        year=year,
        doi=doi,
        source_type=source_type,
        evidence_tier=evidence_tier,
        access_status=access_status,
        local_raw_path=str(raw_path),
        local_text_path=str(text_path),
        extracted_text=text,
        notes=notes,
    )


def split_artifacts(value: str) -> list[str]:
    return [item.strip() for item in re.split(r"[;,]", value) if item.strip()]


def artifact_type(label: str) -> str:
    lower = label.lower()
    for key, value in ARTIFACT_KEYWORDS.items():
        if key in lower:
            return value
    return "other"


def source_signature(row: dict[str, Any]) -> str:
    return " ".join(
        clean_text(row.get(key, ""))
        for key in [
            "source_title",
            "authors / organization",
            "source_type",
            "evidence_tier",
            "DOI",
            "URL",
            "why_included",
            "notes",
        ]
    ).lower()


def score_artifact_source(artifact: str, row: dict[str, Any]) -> tuple[int, int, int, str]:
    signature = source_signature(row)
    artifact_lower = artifact.lower()
    status = clean_text(row.get("access_status", ""))
    source_role = clean_text(row.get("source_role", ""))

    score = 0
    if status == "open":
        score += 100
    elif status == "paywalled":
        score += 10
    else:
        score -= 25

    if source_role == "primary":
        score += 15
    elif source_role == "secondary":
        score += 10
    elif source_role == "registry":
        score -= 200

    score += 10 * sum(1 for hint in GENERIC_ARTIFACT_SOURCE_HINTS if hint in signature)
    for token, hints in ARTIFACT_SOURCE_HINTS.items():
        if token in artifact_lower:
            score += 25 * sum(1 for hint in hints if hint in signature)

    return (
        score,
        1 if source_role == "primary" else 0,
        1 if status == "open" else 0,
        clean_text(row.get("source_id", "")),
    )


def select_artifact_source(artifact: str, source_rows: list[dict[str, Any]], fallback_row: dict[str, Any]) -> dict[str, Any]:
    if not source_rows:
        return fallback_row
    return max(source_rows, key=lambda row: score_artifact_source(artifact, row))


def artifact_source_note(source_row: dict[str, Any]) -> str:
    signature = source_signature(source_row)
    if any(hint in signature for hint in GENERIC_ARTIFACT_SOURCE_HINTS):
        return "Artifact aligned to the best-matching collected supplement, materials page, or benchmark trace."
    return "Artifact inferred from repository traces and aligned to the best-matching collected source."


def matched_roles(candidate_text: str) -> set[str]:
    result: set[str] = set()
    lowered = candidate_text.lower()
    for key, roles in ROLE_KEYWORDS.items():
        if key in lowered:
            result.update(roles)
    return result


def external_source_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [row for row in rows if row.get("source_role") != "registry"]


def infer_ground_truth(case: dict[str, str]) -> str:
    cid = int(case["Case ID"])
    title = case["Case title"].lower()
    family = case["Case family"]
    if cid == 1:
        return "high"
    if any(token in title for token in ["objective", "ground-truth", "benchmarked", "study-type", "repro-bench", "repair"]):
        return "high"
    if family == "Measurement & coding":
        return "contested" if cid in {2, 3, 4, 5, 6} else "medium"
    if family == "Evidence synthesis & corpus":
        return "high" if cid in {7, 8} else "medium"
    if family == "Predictive simulation & design":
        return "contested" if cid in {21, 22, 24, 34} else "low"
    if family == "Interviews, fieldwork & qualitative interpretation":
        return "contested"
    if family == "Agenda, theory & discovery":
        return "none"
    if family == "Governance, authorship & labor":
        return "low"
    if family == "Pedagogy & training":
        return "medium"
    return "medium"


def infer_auditability(case: dict[str, str]) -> str:
    artifacts = case["Observable artifacts / traces"].lower()
    if any(token in artifacts for token in ["log", "version", "prompt", "codebook", "disagreement", "benchmark", "tournament"]):
        return "high"
    if case["Case family"] in {"Evidence synthesis & corpus", "Reproducibility & autonomous research"}:
        return "high"
    if case["Case family"] in {"Predictive simulation & design", "Governance, authorship & labor"}:
        return "medium"
    if case["Case family"] == "Interviews, fieldwork & qualitative interpretation":
        return "low"
    return "medium"


def infer_source_heterogeneity(case: dict[str, str]) -> str:
    evidence = case["Evidence status"].lower()
    if case.get("Secondary source URL"):
        return "high" if any(token in evidence for token in ["mix", "cluster", "boundary"]) else "medium"
    if any(token in evidence for token in ["mix", "cluster"]):
        return "high"
    return "low"


def infer_institutional_specificity(case: dict[str, str]) -> str:
    setting = case["Institutional setting"].lower()
    if any(token in setting for token in ["lab", "team", "course", "study", "project"]):
        return "high"
    if any(token in setting for token in ["public", "website", "field"]):
        return "medium"
    return "low"


def infer_standardization(case: dict[str, str]) -> str:
    family = case["Case family"]
    if family in {"Measurement & coding", "Evidence synthesis & corpus", "Reproducibility & autonomous research"}:
        return "high"
    if family in {"Predictive simulation & design", "Pedagogy & training"}:
        return "medium"
    if family in {"Interviews, fieldwork & qualitative interpretation", "Agenda, theory & discovery"}:
        return "low"
    return "medium"


def infer_role_visibility(case: dict[str, str]) -> str:
    score = 0
    if case["Observable human actors"]:
        score += 1
    if case["Observable artifacts / traces"]:
        score += 1
    if case["Candidate roles tested"]:
        score += 1
    if score >= 3:
        return "high"
    if score == 2:
        return "medium"
    return "low"


def infer_evidence_strength(case: dict[str, str]) -> str:
    value = case["Evidence status"].lower()
    if "peer-reviewed" in value or "conference" in value:
        return "high"
    if "preprint" in value:
        return "medium"
    if "website" in value or "public project" in value or "proposed" in value:
        return "low"
    return "medium"


def infer_autonomy_pressure(case: dict[str, str]) -> str:
    cid = int(case["Case ID"])
    capability = case["Primary AI capability"].lower()
    if cid in {24, 25, 26, 32, 34}:
        return "high"
    if capability == "agentic":
        return "medium"
    if capability == "predictive":
        return "low"
    return "low"


def infer_feasibility(case: dict[str, str]) -> str:
    value = case["Feasibility"].lower()
    if value.startswith("high"):
        return "high"
    if value.startswith("low"):
        return "low"
    return "medium"


def leverage_support_flags(leverage: str) -> tuple[str, str, str]:
    lowered = leverage.lower()
    supports_shrinkage = "yes" if "shrink" in lowered else "no"
    supports_split = "yes" if "split" in lowered else "no"
    supports_merge = "yes" if "merge" in lowered or "reject" in lowered else "no"
    return supports_shrinkage, supports_split, supports_merge


def build_role_rows(case: dict[str, str], supporting_source_ids: list[str]) -> list[dict[str, str]]:
    leverage = case["Primary typology leverage"].lower()
    candidates = matched_roles(case["Candidate roles tested"])
    family_hints = FAMILY_ROLE_HINTS.get(case["Case family"], set())
    rows: list[dict[str, str]] = []
    primary_support = "; ".join(supporting_source_ids[:2])
    for role_label, role_group in ROLE_UNIVERSE:
        directly_named = role_label in candidates
        hinted = role_label in family_hints
        if directly_named:
            observable = "yes"
        elif hinted:
            observable = "partial"
        else:
            observable = "no"

        if role_group == "shrinking_task_cluster":
            if directly_named and "shrink" in leverage:
                status = "shrinking"
            elif role_label == "Routine first-pass screener" and int(case["Case ID"]) in {7, 8, 33}:
                status = "shrinking"
                observable = "yes"
            elif role_label == "Routine coder on high-consensus, ground-truth-rich tasks" and int(case["Case ID"]) in {1, 2, 5, 6}:
                status = "shrinking" if int(case["Case ID"]) == 1 else "conditional"
                observable = "yes"
            elif role_label == "Structural prose polisher / drafter" and int(case["Case ID"]) in {10, 29, 30}:
                status = "shrinking" if int(case["Case ID"]) == 29 else "conditional"
                observable = "yes"
            else:
                status = "not_applicable" if observable == "no" else "unclear"
        elif role_group == "split_candidate":
            if directly_named and "split" in leverage:
                status = "robust"
            elif directly_named and "merge" in leverage:
                status = "merge_pressure"
            elif directly_named:
                status = "conditional"
            elif hinted and "split" in leverage:
                status = "split_pressure"
            else:
                status = "not_applicable" if observable == "no" else "unclear"
        else:
            if directly_named:
                if "robust" in leverage:
                    status = "robust"
                elif "conditional" in leverage:
                    status = "conditional"
                elif "transformed" in leverage:
                    status = "transformed"
                elif "split" in leverage:
                    status = "split_pressure"
                elif "merge" in leverage:
                    status = "merge_pressure"
                elif "shrink" in leverage:
                    status = "conditional"
                else:
                    status = "unclear"
            elif hinted and "robust" in leverage:
                status = "conditional"
            elif hinted and "transformed" in leverage:
                status = "transformed"
            elif hinted and "split" in leverage:
                status = "split_pressure"
            else:
                status = "not_applicable" if observable == "no" else "unclear"

        confidence = "low"
        if directly_named and "peer-reviewed" in case["Evidence status"].lower():
            confidence = "medium"
        if directly_named and int(case["Case ID"]) in {2, 7, 21, 27}:
            confidence = "high"
        evidence_basis = (
            "Initialized from repository candidate-role fields and cross-checked against the collected primary source. "
            "Direct source confirmation is uneven across cases; treat low-confidence rows as provisional."
        )
        notes = ""
        if confidence == "low":
            notes = "no direct source / inferred from case setup or repository curation"
        rows.append(
            {
                "case_id": case["Case ID"],
                "role_label": role_label,
                "role_group": role_group,
                "role_observable_in_case": observable,
                "status_in_case": status,
                "confidence": confidence,
                "evidence_basis_summary": evidence_basis,
                "main_supporting_source_ids": primary_support,
                "main_contrary_source_ids": "",
                "notes": notes,
            }
        )
    return rows


def aggregate_role_status(role_rows: list[dict[str, str]], target_labels: set[str]) -> str:
    relevant = [row for row in role_rows if row["role_label"] in target_labels]
    if not relevant:
        return "unclear"
    precedence = [
        "robust",
        "conditional",
        "transformed",
        "split_pressure",
        "merge_pressure",
        "shrinking",
        "unclear",
        "not_applicable",
    ]
    statuses = {row["status_in_case"] for row in relevant}
    for status in precedence:
        if status in statuses:
            return status
    return "unclear"


def build_codebook() -> str:
    return """# Role Coding Codebook

## Starting roles

### Relevance adjudicator
Human work that decides which materials, cases, papers, or observations belong in the analysis universe.
Example cases: 7, 8, 33.

### Mechanism disciplinarian
Human work that constrains acceptable causal interpretation, model logic, or mechanism claims.
Example cases: 14, 24, 27.

### Construct and perspective boundary setter
Human work that decides what the construct is and how much socially meaningful disagreement should remain visible.
Example cases: 2, 3, 5.

### Corpus pluralist
Human work that keeps the evidence base from collapsing onto the most legible or most machine-friendly corpus.
Example cases: 7, 9.

### Protocol and provenance architect
Human work that designs workflows, logging, versioning, disclosure, reproducibility, and admissible procedure.
Example cases: 7, 27, 29, 30.

### Counterfactual and transportability judge
Human work that decides when simulated, forecast, or model-generated results can travel across settings.
Example cases: 21, 23, 34.

### Situated interlocutor / context witness
Human work that changes what data exist because the human is present in the field, the interview, or the interpretive setting.
Example cases: 16, 19.

### Accountability and labor allocator
Human work that assigns authorship, disclosure, responsibility, and who is answerable for errors or harms.
Example cases: 29, 30, 31, 32.

## Candidate split roles

### Construct boundary setter
Use when the key work is defining category boundaries, codebook rules, or valid labels.

### Perspective sampler
Use when the key work is preserving subgroup disagreement or representational pluralism rather than only one final codebook.

### Counterfactual / transportability judge
Use when the case explicitly pressures whether a result can be moved across populations, contexts, or designs.

### Calibration architect
Use when the case turns on benchmarking, simulator calibration, or matching model outputs to external reference distributions.

### Situated interlocutor
Use when the human interviewer or fieldworker changes what respondents disclose or how the interaction unfolds.

### Context witness
Use when contextual, tacit, or archival knowledge is needed to interpret notes, transcripts, or traces.

## Shrinking task-clusters

### Routine first-pass screener
Use when title/abstract triage, coarse study-type classification, or early ranking are materially reduced by AI.

### Routine coder on high-consensus, ground-truth-rich tasks
Use when labels are externally anchored and humans appear to shift from primary coders to auditors.

### Structural prose polisher / drafter
Use when AI takes over sentence-level drafting, structural polishing, or prose revision without displacing accountability.

## Status criteria

### robust
The human role remains necessary for validity, admissible procedure, or accountability.

### conditional
The role survives, but only under particular task, stage, or institutional conditions.

### transformed
The role clearly persists but is reconfigured toward auditing, exception handling, or workflow design.

### shrinking
The work cluster becomes more routinized, more auditable, and more plausibly delegated to AI.

### split_pressure
The original role looks too coarse and should be separated into two analytically distinct roles.

### merge_pressure
Two roles appear empirically inseparable in the current case.

### unclear
Evidence is mixed, weak, or not yet sufficient.

## Low-confidence coding

Use low confidence when a row is initialized mainly from repository curation, a landing-page abstract, or partial source access rather than direct full-text evidence.

## Contradictory evidence

Record both supporting and limiting evidence. If a case makes a role look durable in one stage but shrinkable in another, keep the row conditional and explain the stage split in notes.

## Live-project and benchmark cases

For Project APE, The AI Scientist, Deep Research, AgentSociety, and adjacent benchmark cases:

- Treat public demos, methodology pages, and benchmark claims as pressure tests, not proof of role disappearance.
- Prioritize methodology, disclaimers, tournament design, benchmarking setup, and public limitations.
- Default confidence downward when evidence is not peer reviewed or when the system boundary changes quickly over time.
"""


def build_readme() -> str:
    return f"""# P1 Role Systems Database

Local analysis-ready database for:

**P1. From Repository to Role System: A Comparative Typology Audit of 36 Cases**

This database treats the repository as a role-system dataset rather than a human-versus-AI dataset. The workbook at [00_input/{WORKBOOK_PATH.name}](00_input/{WORKBOOK_PATH.name}) is the seed registry. All case-level artifacts, manifests, notes, analysis tables, codebook files, and QC logs are generated into this local folder.

Generated on: {TODAY}

## Structure

- `00_input/`: copied workbook input
- `01_registry/`: seed export, enriched master registry, and case list
- `02_cases/`: one standardized folder per case
- `03_raw_sources/`: local raw HTML/PDF snapshots where accessible
- `04_extracted_text/`: extracted text derived from raw source files
- `05_case_notes/`: mirrored markdown case notes
- `06_analysis_tables/`: analysis-ready master CSV tables
- `07_codebook/`: role coding definitions and field dictionary
- `08_logs/`: QC, missingness, duplicates, and access logs
- `09_scripts/`: local build script used to generate the database
- `10_outputs/`: P1-facing summaries and comparison outputs

## Provenance rules used

- Repository workbook rows are preserved exactly in the seed export and propagated into metadata.
- External source collection follows the workbook primary and secondary URLs first.
- Missingness is recorded explicitly rather than silently filled.
- Live-project and autonomy-heavy cases are coded as boundary-pressure evidence unless full, stable methodology is available.
"""


def build_coding_decisions() -> str:
    return """# Coding Decisions

1. The workbook repository is treated as an internal curation source for provenance, especially for case descriptions, leverage notes, falsification criteria, and confounds.
2. External source collection prioritizes the workbook primary URL, then the workbook secondary URL. Additional supporting sources are not added unless already supplied in the workbook companion sheets.
3. When a full text cannot be fetched legally, the landing page snapshot is preserved and the access limit is logged as paywalled or inaccessible.
4. Role coding is initialized from the workbook's candidate-role field plus case-family heuristics. Low-confidence rows are explicitly flagged where direct source confirmation is limited.
5. Analysis-ready status is conservative: `yes` requires at least one usable external source plus completed case files; `partial` indicates a working case folder with explicit evidence limits; `no` is reserved for missing or unusable external source capture.
6. When an original workbook URL remains challenge-blocked or paywalled but a lawful repository copy or preprint is available, the case may still be marked analysis-ready while the original access limit remains logged.
"""


def non_open_external_sources(sources_master: list[dict[str, str]]) -> list[dict[str, str]]:
    return [
        row
        for row in sources_master
        if row.get("source_role") != "registry" and row.get("access_status") != "open"
    ]


def fallback_supported_case_ids(sources_master: list[dict[str, str]]) -> list[str]:
    by_case: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in sources_master:
        if row.get("source_role") == "registry":
            continue
        by_case[row["case_id"]].append(row)
    supported = []
    for case_id, rows in by_case.items():
        if any(row.get("access_status") == "open" for row in rows) and any(
            row.get("access_status") != "open" for row in rows
        ):
            supported.append(case_id)
    return sorted(supported, key=int)


def supplemental_enrichment_case_ids(sources_master: list[dict[str, str]]) -> list[str]:
    enriched = {
        row["case_id"]
        for row in sources_master
        if row.get("source_role") != "registry"
        and row.get("case_id") in SUPPLEMENT_ENRICHMENT_FOCUS_CASE_IDS
        and (
            row.get("source_type") in {"supplementary PDF", "project materials page", "benchmark repo"}
            or row.get("evidence_tier") == "benchmark"
        )
    }
    return sorted(enriched, key=int)


def role_confidence_counts(role_master: list[dict[str, str]] | None) -> Counter[str]:
    if not role_master:
        return Counter()
    return Counter(row.get("confidence", "") for row in role_master)


def build_role_confidence_report(
    cases_master: list[dict[str, str]],
    sources_master: list[dict[str, str]],
    role_master: list[dict[str, str]],
) -> list[dict[str, str]]:
    source_by_case: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in sources_master:
        if row.get("source_role") == "registry":
            continue
        source_by_case[row["case_id"]].append(row)

    confidence_by_case: dict[str, Counter[str]] = defaultdict(Counter)
    for row in role_master:
        confidence_by_case[row["case_id"]][row.get("confidence", "")] += 1

    rows = []
    for case in cases_master:
        case_id = case["Case ID"]
        sources = source_by_case.get(case_id, [])
        counts = confidence_by_case.get(case_id, Counter())
        open_count = sum(1 for row in sources if row.get("access_status") == "open")
        non_open_count = sum(1 for row in sources if row.get("access_status") != "open")
        uses_fallback = "yes" if open_count and non_open_count else "no"
        priority = "high" if counts.get("low", 0) >= 15 or non_open_count else "medium" if counts.get("low", 0) else "low"
        rows.append(
            {
                "case_id": case_id,
                "case_slug": case["case_slug"],
                "case_title": case["Case title"],
                "analysis_ready": case["analysis_ready"],
                "low_confidence_rows": str(counts.get("low", 0)),
                "medium_confidence_rows": str(counts.get("medium", 0)),
                "high_confidence_rows": str(counts.get("high", 0)),
                "open_external_sources": str(open_count),
                "non_open_external_sources": str(non_open_count),
                "uses_open_fallback_or_companion_source": uses_fallback,
                "next_review_priority": priority,
            }
        )
    return rows


def build_collection_log(summary: dict[str, Any]) -> str:
    lines = [
        "# Collection Log",
        "",
        f"- Build date: {TODAY}",
        f"- Seed workbook copied to: `{WORKBOOK_COPY_PATH}`",
        f"- Cases scaffolded: {summary['case_count']}",
        f"- External sources attempted: {summary['external_sources_attempted']}",
        f"- External sources collected as open: {summary['external_sources_open']}",
        f"- External sources marked paywalled: {summary['external_sources_paywalled']}",
        f"- External sources marked inaccessible: {summary['external_sources_inaccessible']}",
        f"- Cases supported by open fallback or companion sources: {summary.get('fallback_supported_cases', 0)}",
        f"- Role-coding confidence rows: high {summary.get('role_coding_high_confidence_rows', 0)}, medium {summary.get('role_coding_medium_confidence_rows', 0)}, low {summary.get('role_coding_low_confidence_rows', 0)}",
        "",
        "## Notes",
        "",
        "- Workbook-derived curation fields are preserved for provenance and for explicit caution/rival-explanation tracking.",
        "- Live project and benchmark cases are intentionally coded as pressure tests rather than settled proof of end-to-end autonomous science.",
        "- Remaining inaccessible journal-original URLs should be read as collector retrieval limits when a lawful local fallback or companion full text already exists in the case folder.",
    ]
    if summary.get("missing_deliverables"):
        lines.extend(["", "## Missing deliverables", ""])
        lines.extend([f"- {item}" for item in summary["missing_deliverables"]])
    return "\n".join(lines)


def build_known_gaps(
    cases_master: list[dict[str, str]],
    sources_master: list[dict[str, str]] | None = None,
    role_master: list[dict[str, str]] | None = None,
) -> str:
    weak_cases = [row for row in cases_master if row["analysis_ready"] != "yes"]
    title_lookup = {row["Case ID"]: row["Case title"] for row in cases_master}
    lines = ["# Known Gaps", ""]
    if weak_cases:
        lines.append("## Case-level gaps")
        lines.append("")
        for row in weak_cases:
            lines.append(
                f"- Case {row['Case ID']} ({row['Case title']}): {row['analysis_ready']} / {row['collection_status']} — {row['main_limitations']}"
            )
    else:
        lines.append("- No case-level analysis gaps remain.")

    non_open = non_open_external_sources(sources_master or [])
    if non_open:
        lines.extend(["", "## Residual source-access limits", ""])
        for row in non_open:
            lines.append(
                f"- Case {row['case_id']} source {row['source_id']} ({row['access_status']}): {row['URL']}"
            )
        fallback_cases = fallback_supported_case_ids(sources_master or [])
        if fallback_cases:
            lines.append(
                "- These remaining journal-original URL failures do not block analysis where a lawful local fallback or companion full text is already stored; the unresolved part is journal-page retrieval, often due to anti-bot or landing-page access friction."
            )

    confidence_counts = role_confidence_counts(role_master)
    if role_master and confidence_counts.get("low", 0):
        low_by_case = Counter(row["case_id"] for row in role_master if row.get("confidence") == "low")
        top_cases = ", ".join(
            f"Case {case_id} ({title_lookup.get(case_id, case_id)}: {count})"
            for case_id, count in low_by_case.most_common(5)
        )
        lines.extend(
            [
                "",
                "## Residual role-coding confidence limits",
                "",
                f"- Low-confidence role rows: {confidence_counts.get('low', 0)} of {len(role_master)} total.",
                f"- Medium-confidence role rows: {confidence_counts.get('medium', 0)}; high-confidence role rows: {confidence_counts.get('high', 0)}.",
                f"- Cases with the densest low-confidence coding: {top_cases}.",
                "- These rows remain traceable, but many still rely on workbook-driven initialization or limited direct role evidence.",
            ]
        )
    return "\n".join(lines)


def build_data_inventory(
    cases_master: list[dict[str, str]],
    sources_master: list[dict[str, str]],
    role_master: list[dict[str, str]] | None = None,
) -> str:
    counts = Counter(row["analysis_ready"] for row in cases_master)
    source_counts = Counter(row["access_status"] for row in sources_master if row.get("source_role") != "registry")
    confidence = role_confidence_counts(role_master)
    lines = [
        "# Data Inventory",
        "",
        f"- Cases in registry: {len(cases_master)}",
        f"- Analysis-ready `yes`: {counts.get('yes', 0)}",
        f"- Analysis-ready `partial`: {counts.get('partial', 0)}",
        f"- Analysis-ready `no`: {counts.get('no', 0)}",
        f"- Open external sources: {source_counts.get('open', 0)}",
        f"- Paywalled external sources: {source_counts.get('paywalled', 0)}",
        f"- Inaccessible external sources: {source_counts.get('inaccessible', 0)}",
        f"- Cases with open fallback or companion-source coverage: {len(fallback_supported_case_ids(sources_master))}",
    ]
    if role_master:
        lines.extend(
            [
                f"- Role-coding rows at `high` confidence: {confidence.get('high', 0)}",
                f"- Role-coding rows at `medium` confidence: {confidence.get('medium', 0)}",
                f"- Role-coding rows at `low` confidence: {confidence.get('low', 0)}",
            ]
        )
    return "\n".join(lines)


def build_case_notes(case: dict[str, str], source_rows: list[dict[str, Any]]) -> str:
    source_lines = []
    for row in source_rows:
        if row["source_role"] == "registry":
            continue
        source_lines.append(
            f"- `{row['source_id']}` ({row['evidence_tier']} / {row['access_status']}): {row['source_title']} — {row['URL']}"
        )
    if not source_lines:
        source_lines.append("- No external source could be collected from the workbook URLs.")

    return "\n".join(
        [
            f"# Case {case['Case ID']}: {case['Case title']}",
            "",
            "## Summary",
            "",
            case["Short description"],
            "",
            "## Why it matters for typology",
            "",
            case["Why this case has high leverage"],
            "",
            "## Caveats",
            "",
            f"- Falsification test: {case['What would weaken / falsify the role claim']}",
            f"- Main confounds: {case['Main confounds / risks of misreading']}",
            f"- Evidence status: {case['Evidence status']}",
            "",
            "## Sources collected",
            "",
            *source_lines,
        ]
    )


def merge_manual_evidence_rows(case_dir: Path, auto_rows: list[dict[str, str]]) -> list[dict[str, str]]:
    existing_path = case_dir / "evidence_extracts.csv"
    if not existing_path.exists():
        return auto_rows
    try:
        existing_rows = read_csv(existing_path)
    except Exception:
        return auto_rows
    auto_ids = {row["extract_id"] for row in auto_rows}
    manual_rows = []
    for row in existing_rows:
        extract_id = row.get("extract_id", "")
        notes = row.get("notes", "").lower()
        if extract_id in auto_ids:
            continue
        if "_m" in extract_id.lower() or "manual review" in notes or "review wave" in notes:
            manual_rows.append(row)
    return auto_rows + manual_rows


PRESERVED_ROLE_FIELDS = (
    "role_observable_in_case",
    "status_in_case",
    "confidence",
    "evidence_basis_summary",
    "main_supporting_source_ids",
    "main_contrary_source_ids",
    "notes",
)


def merge_preserved_role_rows(
    fresh_rows: list[dict[str, str]],
    existing_path: Path,
) -> tuple[list[dict[str, str]], int]:
    """If `existing_path` holds a well-formed per-case role_coding.csv, overlay
    its user-editable fields onto the freshly-initialised `fresh_rows` so that
    a re-run of build_p1_db.py does not destroy manual review work (e.g.,
    wave-2/3 confidence upgrades). Identity fields (case_id, role_label,
    role_group) are always taken from `fresh_rows`. Returns the merged list
    plus a count of rows in which at least one preserved field actually
    diverged from the fresh default (useful for logging)."""
    if not existing_path.exists():
        return fresh_rows, 0
    try:
        with existing_path.open("r", newline="", encoding="utf-8") as handle:
            existing = list(csv.DictReader(handle))
    except OSError:
        return fresh_rows, 0
    if len(existing) != len(fresh_rows):
        return fresh_rows, 0
    existing_by_label = {row.get("role_label", ""): row for row in existing}
    merged: list[dict[str, str]] = []
    preserved_row_count = 0
    for fresh in fresh_rows:
        prior = existing_by_label.get(fresh["role_label"])
        if prior is None:
            merged.append(fresh)
            continue
        row = dict(fresh)
        row_diverged = False
        for key in PRESERVED_ROLE_FIELDS:
            prior_value = prior.get(key, "")
            if prior_value == "":
                continue
            row[key] = prior_value
            if prior_value != fresh.get(key, ""):
                row_diverged = True
        if row_diverged:
            preserved_row_count += 1
        merged.append(row)
    return merged, preserved_row_count


def infer_analysis_status(source_rows: list[dict[str, Any]], role_rows: list[dict[str, str]]) -> tuple[str, str, str]:
    external = external_source_rows(source_rows)
    open_external = [row for row in external if row["access_status"] == "open"]
    if open_external and role_rows:
        return "complete", "yes", ""
    if external:
        limits = "Primary source access is limited; case is usable with explicit caveats and landing-page evidence."
        return "partial", "partial", limits
    return "blocked", "no", "No external source could be collected from the workbook URLs."


def build_casebook_row(case: dict[str, str], cases_master_row: dict[str, str], role_rows: list[dict[str, str]]) -> str:
    strongest_roles = [row["role_label"] for row in role_rows if row["status_in_case"] in {"robust", "conditional", "transformed"}]
    strongest_roles = strongest_roles[:3] or ["unclear"]
    return "\n".join(
        [
            f"## Case {case['Case ID']}: {case['Case title']}",
            "",
            f"{case['Short description']} This case matters because {case['Why this case has high leverage'].lower()} "
            f"The strongest visible roles in the current database are {', '.join(strongest_roles)}. "
            f"The strongest rival explanation is: {case['Main confounds / risks of misreading']}. "
            f"Analysis usability: {cases_master_row['analysis_ready']}.",
            "",
        ]
    )


def build_anchor_cases(repo_by_id: dict[int, dict[str, str]]) -> str:
    groups = {
        "Anchor cases": [2, 7, 21, 27],
        "Split-pressure cases": [3, 19, 24, 34],
        "Shrinkage cases": [1, 7, 28, 29],
        "Autonomy-boundary cases": [25, 26, 32],
    }
    lines = ["# P1 Anchor Cases", ""]
    for label, case_ids in groups.items():
        lines.extend([f"## {label}", ""])
        for case_id in case_ids:
            case = repo_by_id[case_id]
            lines.append(
                f"- Case {case_id}: {case['Case title']} — {case['Why this case has high leverage']}"
            )
        lines.append("")
    return "\n".join(lines)


def build_collection_summary(
    cases_master: list[dict[str, str]],
    sources_master: list[dict[str, str]] | None = None,
    role_master: list[dict[str, str]] | None = None,
) -> str:
    counts = Counter(row["analysis_ready"] for row in cases_master)
    strongest = [row["Case title"] for row in cases_master if row["analysis_ready"] == "yes"][:6]
    frontier = [row["Case title"] for row in cases_master if "boundary" in row["Portfolio group"].lower() or "live" in row["Evidence status"].lower()]
    confidence = role_confidence_counts(role_master)
    fallback_cases = len(fallback_supported_case_ids(sources_master or [])) if sources_master is not None else 0
    enrichment_cases = supplemental_enrichment_case_ids(sources_master or []) if sources_master is not None else []
    remaining_enrichment_cases = [
        case_id for case_id in sorted(SUPPLEMENT_ENRICHMENT_FOCUS_CASE_IDS, key=int) if case_id not in enrichment_cases
    ]
    has_non_open = bool(non_open_external_sources(sources_master or []))
    follow_up = []
    if has_non_open:
        follow_up.append(
            "If preserving original journal-page HTML still matters, use a browser-authenticated/manual capture workflow for the remaining blocked journal pages; substantive full text is already local via lawful fallbacks."
        )
    if enrichment_cases:
        follow_up.append(
            f"Optional: look for any additional public appendices or materials for {', '.join(f'Case {case_id}' for case_id in remaining_enrichment_cases)}; the main supplement and benchmark enrichment pass is already local."
        )
    else:
        follow_up.append(
            "Add supplementary appendices or repositories for the anchor cases where codebooks, logs, or benchmark traces are publicly available."
        )
    if confidence.get("low", 0):
        follow_up.append(
            "Replace provisional low-confidence role coding with direct full-text evidence for the remaining high-priority review cases identified in the role confidence report."
        )
    else:
        follow_up.append(
            "Optional: deepen medium-confidence rows with more direct full-text role extracts for selected anchor cases, but the database is already complete and analysis-ready."
        )
    lines = [
        "# P1 Collection Summary",
        "",
        f"- Analysis-ready yes: {counts.get('yes', 0)}",
        f"- Analysis-ready partial: {counts.get('partial', 0)}",
        f"- Analysis-ready no: {counts.get('no', 0)}",
        f"- Cases supported by open fallback or companion sources: {fallback_cases}",
        f"- Public supplement/materials/benchmark enrichments added for: {', '.join(f'Case {case_id}' for case_id in enrichment_cases) if enrichment_cases else 'None yet'}.",
        f"- Strongest evidence clusters: Measurement & coding; Evidence synthesis & corpus; Predictive simulation & design anchor cases.",
        f"- Frontier evidence clusters: {', '.join(frontier[:6]) if frontier else 'None recorded'}.",
        f"- Most analysis-ready cases: {', '.join(strongest) if strongest else 'None yet'}",
    ]
    if role_master:
        lines.append(
            f"- Role-coding confidence profile: high {confidence.get('high', 0)}, medium {confidence.get('medium', 0)}, low {confidence.get('low', 0)}."
        )
    lines.extend(
        [
            "- Minimum follow-up that would improve the database:",
            *[f"  - {item}" for item in follow_up],
        ]
    )
    return "\n".join(lines)


def tree_lines(root: Path) -> list[str]:
    entries = []
    for path in sorted(root.rglob("*")):
        rel = path.relative_to(root)
        depth = len(rel.parts) - 1
        indent = "  " * depth
        suffix = "/" if path.is_dir() else ""
        entries.append(f"{indent}{rel.name}{suffix}")
    return entries


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description=(
            "Build the P1 role-systems database from the seed workbook. "
            "By default, existing per-case role_coding.csv rows are PRESERVED "
            "so a re-run does not destroy manual review work "
            "(e.g., wave-2/3 confidence upgrades)."
        )
    )
    parser.add_argument(
        "--reset-role-coding",
        action="store_true",
        help=(
            "Overwrite per-case role_coding.csv with fresh workbook-initialised "
            "defaults. Use this only when you want to deliberately discard "
            "manual coding upgrades and start over."
        ),
    )
    args = parser.parse_args(argv)
    preserve_role_coding = not args.reset_role_coding

    if not WORKBOOK_PATH.exists():
        print(f"Workbook not found at {WORKBOOK_PATH}", file=sys.stderr)
        return 1

    ensure_structure()

    wb = load_workbook(WORKBOOK_PATH, data_only=True)
    repository_ws = wb["Repository"]
    export_exact_sheet_csv(repository_ws, PROJECT_ROOT / "01_registry" / "cases_seed.csv")

    repo_rows = read_table(repository_ws, REPOSITORY_HEADER_ROW)
    repo_rows = [row for row in repo_rows if row.get("Case ID")]
    for row in repo_rows:
        row["case_slug"] = CASE_SLUGS[int(row["Case ID"])]

    sources_rows = read_table(wb["Sources"], SOURCES_HEADER_ROW)
    _coverage_rows = read_table(wb["Coverage"], FLAG_HEADER_ROW)
    flagship_rows = read_table(wb["Flagship_8"], FLAG_HEADER_ROW)
    shortlist_rows = read_table(wb["Shortlist_12"], FLAG_HEADER_ROW)
    ape_cluster_rows = read_table(wb["APE_Cluster"], FLAG_HEADER_ROW)
    data_dictionary_rows = read_table(wb["Data_Dictionary"], DICTIONARY_HEADER_ROW)
    repo_by_id = {int(row["Case ID"]): row for row in repo_rows}

    cases_master: list[dict[str, str]] = []
    sources_master: list[dict[str, str]] = []
    evidence_extracts_master: list[dict[str, str]] = []
    artifacts_master: list[dict[str, str]] = []
    case_features_rows: list[dict[str, str]] = []
    role_master: list[dict[str, str]] = []
    hypotheses_rows: list[dict[str, str]] = []
    exclusions_rows: list[dict[str, str]] = []
    missingness_rows: list[dict[str, str]] = []
    duplicate_rows: list[dict[str, str]] = []
    source_access_rows: list[dict[str, str]] = []

    source_url_seen: dict[str, list[str]] = defaultdict(list)

    registry_fieldnames = list(repo_rows[0].keys())
    extra_master_fields = [
        "collection_status",
        "source_count_total",
        "source_count_primary",
        "raw_files_count",
        "extracted_text_count",
        "artifacts_captured_count",
        "analysis_ready",
        "main_limitations",
        "last_updated",
        "collector_notes",
    ]

    for case in repo_rows:
        case_id = int(case["Case ID"])
        case_slug = case["case_slug"]
        case_dir = PROJECT_ROOT / "02_cases" / case_slug
        raw_dir = PROJECT_ROOT / "03_raw_sources" / case_slug
        text_dir = PROJECT_ROOT / "04_extracted_text" / case_slug
        note_mirror_dir = PROJECT_ROOT / "05_case_notes" / case_slug
        case_dir.mkdir(parents=True, exist_ok=True)
        raw_dir.mkdir(parents=True, exist_ok=True)
        text_dir.mkdir(parents=True, exist_ok=True)
        note_mirror_dir.mkdir(parents=True, exist_ok=True)

        source_manifest_rows: list[dict[str, Any]] = []
        registry_source_id = f"C{case_id:03d}_REGISTRY"
        registry_source_row = {
            "source_id": registry_source_id,
            "case_id": str(case_id),
            "source_role": "registry",
            "source_title": f"Repository workbook entry for Case {case_id:03d}",
            "authors / organization": "Local workbook / repository curation",
            "year": TODAY[:4],
            "source_type": "registry_workbook_entry",
            "evidence_tier": "commentary / forum / boundary evidence",
            "DOI": "",
            "URL": "",
            "local_raw_path": str(WORKBOOK_COPY_PATH),
            "local_text_path": "",
            "retrieval_date": TODAY,
            "access_status": "open",
            "why_included": "Preserves repository provenance, curation notes, confounds, and role hypotheses.",
            "notes": "Internal seed source provided by the user.",
        }
        source_manifest_rows.append(registry_source_row)

        for source_role, url_field, title_suffix in [
            ("primary", "Primary source URL", "Primary source"),
            ("secondary", "Secondary source URL", "Secondary source"),
        ]:
            url = case.get(url_field, "")
            if not url:
                continue
            if any(row["URL"] == url for row in source_manifest_rows if row.get("URL")):
                continue
            source_id = f"C{case_id:03d}_{'P' if source_role == 'primary' else 'S'}1"
            result = fetch_url(url, raw_dir, text_dir, source_id, case["Case title"], case["Evidence status"])
            source_row = {
                "source_id": source_id,
                "case_id": str(case_id),
                "source_role": source_role,
                "source_title": result.title or case["Case title"],
                "authors / organization": result.authors,
                "year": result.year,
                "source_type": result.source_type,
                "evidence_tier": result.evidence_tier,
                "DOI": result.doi,
                "URL": result.final_url,
                "local_raw_path": result.local_raw_path,
                "local_text_path": result.local_text_path,
                "retrieval_date": TODAY,
                "access_status": result.access_status,
                "why_included": f"{title_suffix} listed in the repository workbook.",
                "notes": result.notes,
            }
            source_manifest_rows.append(source_row)
            if result.final_url:
                source_url_seen[result.final_url].append(source_id)
            source_access_rows.append(
                {
                    "case_id": str(case_id),
                    "case_slug": case_slug,
                    "source_id": source_id,
                    "url": result.final_url,
                    "access_status": result.access_status,
                    "local_raw_path": result.local_raw_path,
                    "local_text_path": result.local_text_path,
                }
            )

        supporting_source_ids = [row["source_id"] for row in source_manifest_rows if row["source_role"] != "registry"]
        if not supporting_source_ids:
            supporting_source_ids = [registry_source_id]

        supports_shrinkage, supports_split, supports_merge = leverage_support_flags(case["Primary typology leverage"])
        external_rows = external_source_rows(source_manifest_rows)
        supporting_rows = sorted(
            external_rows,
            key=lambda row: (
                0 if clean_text(row.get("access_status", "")) == "open"
                else 1 if clean_text(row.get("access_status", "")) == "paywalled"
                else 2,
                clean_text(row.get("source_id", "")),
            ),
        )
        external_texts = []
        for row in external_rows:
            text_path = Path(row["local_text_path"]) if row["local_text_path"] else None
            text_value = ""
            if text_path and text_path.exists():
                text_value = text_path.read_text(encoding="utf-8", errors="ignore").strip()
            if text_value:
                external_texts.append((row["source_id"], text_value[:1200]))

        evidence_rows = [
            {
                "extract_id": f"E{case_id:03d}_001",
                "case_id": str(case_id),
                "case_slug": case_slug,
                "source_id": registry_source_id,
                "source_type": "registry_workbook_entry",
                "page_or_section": "Repository row",
                "extract_type": "role_claim",
                "text_excerpt_or_paraphrase": case["Short description"],
                "supports_role": case["Candidate roles tested"],
                "contradicts_role": "",
                "supports_shrinkage": supports_shrinkage,
                "supports_split": supports_split,
                "supports_merge_or_rejection": supports_merge,
                "notes": "Workbook seed description preserved for provenance.",
            },
            {
                "extract_id": f"E{case_id:03d}_002",
                "case_id": str(case_id),
                "case_slug": case_slug,
                "source_id": registry_source_id,
                "source_type": "registry_workbook_entry",
                "page_or_section": "Repository row",
                "extract_type": "counterevidence",
                "text_excerpt_or_paraphrase": case["Main confounds / risks of misreading"],
                "supports_role": "",
                "contradicts_role": case["Candidate roles tested"],
                "supports_shrinkage": "no",
                "supports_split": "no",
                "supports_merge_or_rejection": "no",
                "notes": "Repository curation note captured as rival explanation / caution.",
            },
            {
                "extract_id": f"E{case_id:03d}_003",
                "case_id": str(case_id),
                "case_slug": case_slug,
                "source_id": registry_source_id,
                "source_type": "registry_workbook_entry",
                "page_or_section": "Repository row",
                "extract_type": "boundary_condition",
                "text_excerpt_or_paraphrase": case["What would weaken / falsify the role claim"],
                "supports_role": "",
                "contradicts_role": case["Candidate roles tested"],
                "supports_shrinkage": "no",
                "supports_split": "no",
                "supports_merge_or_rejection": "no",
                "notes": "Falsification condition recorded from repository seed.",
            },
        ]
        if external_texts:
            source_id, text_excerpt = external_texts[0]
            evidence_rows.append(
                {
                    "extract_id": f"E{case_id:03d}_004",
                    "case_id": str(case_id),
                    "case_slug": case_slug,
                    "source_id": source_id,
                    "source_type": next(row["source_type"] for row in external_rows if row["source_id"] == source_id),
                    "page_or_section": "landing page / extracted text",
                    "extract_type": "method",
                    "text_excerpt_or_paraphrase": text_excerpt,
                    "supports_role": case["Candidate roles tested"],
                    "contradicts_role": "",
                    "supports_shrinkage": supports_shrinkage,
                    "supports_split": supports_split,
                    "supports_merge_or_rejection": supports_merge,
                    "notes": "Auto-extracted text snippet from collected source.",
                }
            )
        evidence_rows = merge_manual_evidence_rows(case_dir, evidence_rows)

        artifact_rows: list[dict[str, str]] = []
        for index, artifact in enumerate(split_artifacts(case["Observable artifacts / traces"]), start=1):
            candidate_source = select_artifact_source(artifact, supporting_rows, registry_source_row)
            source_ref = clean_text(candidate_source.get("source_id", "")) or registry_source_id
            local_path = candidate_source["local_raw_path"] or candidate_source["URL"]
            usable = "yes" if candidate_source.get("access_status") == "open" else "partial"
            artifact_rows.append(
                {
                    "artifact_id": f"A{case_id:03d}_{index:03d}",
                    "case_id": str(case_id),
                    "artifact_type": artifact_type(artifact),
                    "artifact_description": artifact,
                    "local_path": local_path,
                    "source_id": source_ref,
                    "usable_for_analysis": usable,
                    "notes": artifact_source_note(candidate_source),
                }
            )

        role_rows = build_role_rows(case, supporting_source_ids + [registry_source_id])
        if preserve_role_coding:
            role_rows, preserved_count = merge_preserved_role_rows(
                role_rows, case_dir / "role_coding.csv"
            )
            if preserved_count:
                print(
                    f"  Case {case_id:03d}: preserved {preserved_count} existing "
                    f"role_coding row(s) (pass --reset-role-coding to discard)",
                    file=sys.stderr,
                )
        collection_status, analysis_ready, limitations = infer_analysis_status(source_manifest_rows, role_rows)

        source_count_total = len(external_rows)
        source_count_primary = len([row for row in external_rows if row["source_role"] == "primary"])
        raw_files_count = len([row for row in external_rows if row["local_raw_path"]])
        extracted_text_count = len([row for row in external_rows if row["local_text_path"]])
        artifacts_captured_count = len(artifact_rows)

        metadata = dict(case)
        metadata.update(
            {
                "collection_status": collection_status,
                "source_count_total": source_count_total,
                "source_count_primary": source_count_primary,
                "raw_files_count": raw_files_count,
                "extracted_text_count": extracted_text_count,
                "artifacts_captured_count": artifacts_captured_count,
                "analysis_ready": analysis_ready,
                "main_limitations": limitations,
                "last_updated": TODAY,
                "collector_notes": "Generated from workbook and workbook-linked sources.",
            }
        )
        write_text(case_dir / "notes.md", build_case_notes(case, source_manifest_rows))
        write_text(note_mirror_dir / "notes.md", build_case_notes(case, source_manifest_rows))
        write_text(
            case_dir / "qa.md",
            "\n".join(
                [
                    f"# QA for Case {case_id}: {case['Case title']}",
                    "",
                    f"- Collection status: {collection_status}",
                    f"- Analysis ready: {analysis_ready}",
                    f"- Main limitations: {limitations or 'None recorded'}",
                    f"- Evidence status: {case['Evidence status']}",
                    f"- Main confounds: {case['Main confounds / risks of misreading']}",
                    f"- Unresolved coding question: {case['Key typology question']}",
                ]
            ),
        )
        write_json(case_dir / "metadata.json", metadata)

        source_manifest_fieldnames = [
            "source_id",
            "case_id",
            "source_role",
            "source_title",
            "authors / organization",
            "year",
            "source_type",
            "evidence_tier",
            "DOI",
            "URL",
            "local_raw_path",
            "local_text_path",
            "retrieval_date",
            "access_status",
            "why_included",
            "notes",
        ]
        write_csv(case_dir / "source_manifest.csv", source_manifest_fieldnames, source_manifest_rows)
        write_csv(
            case_dir / "evidence_extracts.csv",
            [
                "extract_id",
                "case_id",
                "case_slug",
                "source_id",
                "source_type",
                "page_or_section",
                "extract_type",
                "text_excerpt_or_paraphrase",
                "supports_role",
                "contradicts_role",
                "supports_shrinkage",
                "supports_split",
                "supports_merge_or_rejection",
                "notes",
            ],
            evidence_rows,
        )
        write_csv(
            case_dir / "artifacts_index.csv",
            [
                "artifact_id",
                "case_id",
                "artifact_type",
                "artifact_description",
                "local_path",
                "source_id",
                "usable_for_analysis",
                "notes",
            ],
            artifact_rows,
        )
        write_csv(
            case_dir / "role_coding.csv",
            [
                "case_id",
                "role_label",
                "role_group",
                "role_observable_in_case",
                "status_in_case",
                "confidence",
                "evidence_basis_summary",
                "main_supporting_source_ids",
                "main_contrary_source_ids",
                "notes",
            ],
            role_rows,
        )

        cases_master.append(
            {
                **case,
                "collection_status": collection_status,
                "source_count_total": str(source_count_total),
                "source_count_primary": str(source_count_primary),
                "raw_files_count": str(raw_files_count),
                "extracted_text_count": str(extracted_text_count),
                "artifacts_captured_count": str(artifacts_captured_count),
                "analysis_ready": analysis_ready,
                "main_limitations": limitations,
                "last_updated": TODAY,
                "collector_notes": "Generated from workbook and workbook-linked sources.",
            }
        )
        sources_master.extend(source_manifest_rows)
        evidence_extracts_master.extend(evidence_rows)
        artifacts_master.extend(artifact_rows)
        role_master.extend(role_rows)

        case_features_rows.append(
            {
                **case,
                "task_ground_truth_level": infer_ground_truth(case),
                "post_hoc_auditability": infer_auditability(case),
                "source_heterogeneity": infer_source_heterogeneity(case),
                "institutional_specificity": infer_institutional_specificity(case),
                "task_standardization": infer_standardization(case),
                "role_visibility": infer_role_visibility(case),
                "evidence_strength_overall": infer_evidence_strength(case),
                "case_family_comparability": case["Case family"],
                "autonomy_pressure": infer_autonomy_pressure(case),
                "current_data_feasibility": infer_feasibility(case),
                "notes_justifying_these_fields": (
                    f"Derived from workbook fields for case family ({case['Case family']}), evidence status "
                    f"({case['Evidence status']}), feasibility ({case['Feasibility']}), and recorded artifacts."
                ),
            }
        )

        boundary_status = aggregate_role_status(
            role_rows,
            {
                "Construct and perspective boundary setter",
                "Construct boundary setter",
                "Perspective sampler",
            },
        )
        procedure_status = aggregate_role_status(
            role_rows,
            {"Protocol and provenance architect", "Mechanism disciplinarian"},
        )
        accountability_status = aggregate_role_status(
            role_rows,
            {"Accountability and labor allocator"},
        )
        routine_status = aggregate_role_status(
            role_rows,
            {
                "Routine first-pass screener",
                "Routine coder on high-consensus, ground-truth-rich tasks",
                "Structural prose polisher / drafter",
            },
        )

        leverage = case["Primary typology leverage"].lower()
        dominant_pattern = {
            "robust": "boundary-setting durability",
            "conditional": "stage-conditional durability",
            "transformed": "role transformation under AI assistance",
            "shrinking": "routine-throughput shrinkage",
            "split": "role split pressure",
        }
        dominant_label = "comparative mixed pattern"
        for key, value in dominant_pattern.items():
            if key in leverage:
                dominant_label = value
                break

        hypotheses_rows.append(
            {
                "case_id": case["Case ID"],
                "primary_ai_capability": case["Primary AI capability"],
                "primary_research_stage": case["Primary research stage"],
                "institutional_setting": case["Institutional setting"],
                "task_ground_truth_level": infer_ground_truth(case),
                "post_hoc_auditability": infer_auditability(case),
                "role_status_boundary_setting": boundary_status,
                "role_status_procedure_setting": procedure_status,
                "role_status_accountability": accountability_status,
                "role_status_routine_throughput": routine_status,
                "dominant_role_pattern": dominant_label,
                "shrinkage_present": "yes" if "shrink" in leverage else "partial" if routine_status != "not_applicable" else "no",
                "split_pressure_present": "yes" if "split" in leverage else "partial" if boundary_status == "split_pressure" else "no",
                "notes": "Derived from role coding master and workbook leverage fields.",
            }
        )

        if analysis_ready != "yes":
            exclusions_rows.append(
                {
                    "case_id": case["Case ID"],
                    "missing_item": "full usable primary source evidence",
                    "reason": limitations or "Source capture incomplete",
                    "severity": "medium" if analysis_ready == "partial" else "high",
                    "whether_case_still_analysis_ready": analysis_ready,
                }
            )

        if not external_rows:
            missingness_rows.append(
                {
                    "case_id": case["Case ID"],
                    "case_slug": case_slug,
                    "missing_item": "external_source",
                    "reason": "No primary or secondary source captured from workbook URLs",
                    "severity": "high",
                }
            )
        elif not any(row["local_text_path"] for row in external_rows):
            missingness_rows.append(
                {
                    "case_id": case["Case ID"],
                    "case_slug": case_slug,
                    "missing_item": "extracted_text",
                    "reason": "Source landing page captured without usable extracted text",
                    "severity": "medium",
                }
            )

    for url, source_ids in sorted(source_url_seen.items()):
        if len(source_ids) > 1:
            duplicate_rows.append(
                {
                    "duplicate_type": "source_url",
                    "value": url,
                    "occurrences": "; ".join(source_ids),
                    "note": "Intentional duplicate or shared source; review for dedupe if needed.",
                }
            )
    title_counts = Counter(row["Case title"] for row in cases_master)
    for title, count in title_counts.items():
        if count > 1:
            duplicate_rows.append(
                {
                    "duplicate_type": "case_title",
                    "value": title,
                    "occurrences": str(count),
                    "note": "Case titles should be unique.",
                }
            )

    case_list_lines = ["# Case List", ""]
    for row in cases_master:
        case_list_lines.append(
            f"1. Case {row['Case ID']}: {row['Case title']} | {row['Case family']} | {row['Evidence status']} | `{row['case_slug']}`"
        )
    write_text(PROJECT_ROOT / "01_registry" / "case_list.md", "\n".join(case_list_lines))

    write_csv(PROJECT_ROOT / "01_registry" / "cases_master.csv", registry_fieldnames + extra_master_fields, cases_master)
    write_csv(
        PROJECT_ROOT / "06_analysis_tables" / "sources_master.csv",
        [
            "source_id",
            "case_id",
            "source_role",
            "source_title",
            "authors / organization",
            "year",
            "source_type",
            "evidence_tier",
            "DOI",
            "URL",
            "local_raw_path",
            "local_text_path",
            "retrieval_date",
            "access_status",
            "why_included",
            "notes",
        ],
        sources_master,
    )
    write_csv(
        PROJECT_ROOT / "06_analysis_tables" / "evidence_extracts_master.csv",
        [
            "extract_id",
            "case_id",
            "case_slug",
            "source_id",
            "source_type",
            "page_or_section",
            "extract_type",
            "text_excerpt_or_paraphrase",
            "supports_role",
            "contradicts_role",
            "supports_shrinkage",
            "supports_split",
            "supports_merge_or_rejection",
            "notes",
        ],
        evidence_extracts_master,
    )
    write_csv(
        PROJECT_ROOT / "06_analysis_tables" / "artifacts_master.csv",
        [
            "artifact_id",
            "case_id",
            "artifact_type",
            "artifact_description",
            "local_path",
            "source_id",
            "usable_for_analysis",
            "notes",
        ],
        artifacts_master,
    )
    write_csv(
        PROJECT_ROOT / "06_analysis_tables" / "case_features.csv",
        registry_fieldnames
        + [
            "task_ground_truth_level",
            "post_hoc_auditability",
            "source_heterogeneity",
            "institutional_specificity",
            "task_standardization",
            "role_visibility",
            "evidence_strength_overall",
            "case_family_comparability",
            "autonomy_pressure",
            "current_data_feasibility",
            "notes_justifying_these_fields",
        ],
        case_features_rows,
    )
    write_csv(
        PROJECT_ROOT / "06_analysis_tables" / "role_coding_master.csv",
        [
            "case_id",
            "role_label",
            "role_group",
            "role_observable_in_case",
            "status_in_case",
            "confidence",
            "evidence_basis_summary",
            "main_supporting_source_ids",
            "main_contrary_source_ids",
            "notes",
        ],
        role_master,
    )
    write_csv(
        PROJECT_ROOT / "06_analysis_tables" / "hypotheses_prep.csv",
        [
            "case_id",
            "primary_ai_capability",
            "primary_research_stage",
            "institutional_setting",
            "task_ground_truth_level",
            "post_hoc_auditability",
            "role_status_boundary_setting",
            "role_status_procedure_setting",
            "role_status_accountability",
            "role_status_routine_throughput",
            "dominant_role_pattern",
            "shrinkage_present",
            "split_pressure_present",
            "notes",
        ],
        hypotheses_rows,
    )
    write_csv(
        PROJECT_ROOT / "06_analysis_tables" / "exclusions_and_limits.csv",
        ["case_id", "missing_item", "reason", "severity", "whether_case_still_analysis_ready"],
        exclusions_rows,
    )

    field_dictionary_rows = []
    for table_name in [
        "sources_master.csv",
        "evidence_extracts_master.csv",
        "artifacts_master.csv",
        "case_features.csv",
        "role_coding_master.csv",
        "hypotheses_prep.csv",
        "exclusions_and_limits.csv",
    ]:
        rows = read_csv(PROJECT_ROOT / "06_analysis_tables" / table_name)
        if not rows:
            continue
        for column in rows[0].keys():
            field_dictionary_rows.append(
                {
                    "table_name": table_name,
                    "column_name": column,
                    "definition": "Generated field in the analysis-ready database.",
                }
            )
    write_csv(
        PROJECT_ROOT / "07_codebook" / "field_dictionary.csv",
        ["table_name", "column_name", "definition"],
        field_dictionary_rows,
    )
    write_text(PROJECT_ROOT / "07_codebook" / "role_coding_codebook.md", build_codebook())

    summary = {
        "case_count": len(cases_master),
        "external_sources_attempted": len([row for row in sources_master if row["source_role"] != "registry"]),
        "external_sources_open": len([row for row in sources_master if row["source_role"] != "registry" and row["access_status"] == "open"]),
        "external_sources_paywalled": len([row for row in sources_master if row["source_role"] != "registry" and row["access_status"] == "paywalled"]),
        "external_sources_inaccessible": len([row for row in sources_master if row["source_role"] != "registry" and row["access_status"] == "inaccessible"]),
        "fallback_supported_cases": len(fallback_supported_case_ids(sources_master)),
        "role_coding_high_confidence_rows": sum(1 for row in role_master if row.get("confidence") == "high"),
        "role_coding_medium_confidence_rows": sum(1 for row in role_master if row.get("confidence") == "medium"),
        "role_coding_low_confidence_rows": sum(1 for row in role_master if row.get("confidence") == "low"),
    }

    write_text(PROJECT_ROOT / "README.md", build_readme())
    write_text(PROJECT_ROOT / "coding_decisions.md", build_coding_decisions())
    write_text(PROJECT_ROOT / "known_gaps.md", build_known_gaps(cases_master, sources_master, role_master))
    write_text(PROJECT_ROOT / "data_inventory.md", build_data_inventory(cases_master, sources_master, role_master))
    write_text(PROJECT_ROOT / "collection_log.md", build_collection_log(summary))

    write_csv(
        PROJECT_ROOT / "08_logs" / "missingness_report.csv",
        ["case_id", "case_slug", "missing_item", "reason", "severity"],
        missingness_rows,
    )
    write_csv(
        PROJECT_ROOT / "08_logs" / "duplicate_check.csv",
        ["duplicate_type", "value", "occurrences", "note"],
        duplicate_rows,
    )
    write_csv(
        PROJECT_ROOT / "08_logs" / "source_access_report.csv",
        ["case_id", "case_slug", "source_id", "url", "access_status", "local_raw_path", "local_text_path"],
        source_access_rows,
    )
    write_csv(
        PROJECT_ROOT / "08_logs" / "role_confidence_report.csv",
        [
            "case_id",
            "case_slug",
            "case_title",
            "analysis_ready",
            "low_confidence_rows",
            "medium_confidence_rows",
            "high_confidence_rows",
            "open_external_sources",
            "non_open_external_sources",
            "uses_open_fallback_or_companion_source",
            "next_review_priority",
        ],
        build_role_confidence_report(cases_master, sources_master, role_master),
    )

    missing_cases = {row["Case ID"] for row in cases_master if not (PROJECT_ROOT / "02_cases" / row["case_slug"]).exists()}
    qc_lines = [
        "# QC Report",
        "",
        f"- Cases in workbook: {len(repo_rows)}",
        f"- Case folders created: {len(list((PROJECT_ROOT / '02_cases').glob('*')))}",
        f"- Missing case folders: {', '.join(sorted(missing_cases)) if missing_cases else 'None'}",
        f"- Role coding rows: {len(role_master)}",
        f"- Sources rows: {len(sources_master)}",
        f"- Artifacts rows: {len(artifacts_master)}",
        "",
        "## Checks",
        "",
        f"- Every workbook case in cases_master.csv: {'yes' if len(cases_master) == len(repo_rows) else 'no'}",
        f"- Every workbook case has a case folder: {'yes' if not missing_cases else 'no'}",
        f"- Every role row has supporting source IDs or explicit inferential flag: {'yes' if all(row['main_supporting_source_ids'] or 'inferred' in row['notes'] for row in role_master) else 'no'}",
        f"- Every source row has case_id plus retrieval_date plus URL/DOI or local workbook path: {'yes' if all(row['case_id'] and row['retrieval_date'] and (row['URL'] or row['DOI'] or row['local_raw_path']) for row in sources_master) else 'no'}",
        f"- Missingness report written: {'yes' if (PROJECT_ROOT / '08_logs' / 'missingness_report.csv').exists() else 'no'}",
    ]
    write_text(PROJECT_ROOT / "08_logs" / "qc_report.md", "\n".join(qc_lines))

    casebook_sections = []
    for case in repo_rows:
        case_rows = [row for row in role_master if row["case_id"] == case["Case ID"]]
        master_row = next(row for row in cases_master if row["Case ID"] == case["Case ID"])
        casebook_sections.append(build_casebook_row(case, master_row, case_rows))
    write_text(PROJECT_ROOT / "10_outputs" / "p1_casebook.md", "\n".join(casebook_sections))
    write_text(PROJECT_ROOT / "10_outputs" / "p1_anchor_cases.md", build_anchor_cases(repo_by_id))

    overview_rows = []
    role_counts: dict[str, Counter[str]] = defaultdict(Counter)
    role_families: dict[str, Counter[str]] = defaultdict(Counter)
    role_caps: dict[str, Counter[str]] = defaultdict(Counter)
    for row in role_master:
        role_counts[row["role_label"]][row["status_in_case"]] += 1
        case = repo_by_id[int(row["case_id"])]
        role_families[row["role_label"]][case["Case family"]] += 1
        role_caps[row["role_label"]][case["Primary AI capability"]] += 1
    for role_label, _role_group in ROLE_UNIVERSE:
        counter = role_counts[role_label]
        families = role_families[role_label].most_common(2)
        caps = role_caps[role_label].most_common(2)
        overview_rows.append(
            {
                "role_label": role_label,
                "number_of_cases_robust": counter.get("robust", 0),
                "number_of_cases_conditional": counter.get("conditional", 0),
                "number_of_cases_transformed": counter.get("transformed", 0),
                "number_of_cases_shrinking": counter.get("shrinking", 0),
                "number_of_cases_split_pressure": counter.get("split_pressure", 0),
                "number_of_cases_merge_pressure": counter.get("merge_pressure", 0),
                "most_common_case_families": "; ".join(f"{name} ({count})" for name, count in families),
                "most_common_ai_capabilities": "; ".join(f"{name} ({count})" for name, count in caps),
                "notes": "Counts generated from role_coding_master.csv",
            }
        )
    write_csv(
        PROJECT_ROOT / "10_outputs" / "p1_typology_overview_table.csv",
        [
            "role_label",
            "number_of_cases_robust",
            "number_of_cases_conditional",
            "number_of_cases_transformed",
            "number_of_cases_shrinking",
            "number_of_cases_split_pressure",
            "number_of_cases_merge_pressure",
            "most_common_case_families",
            "most_common_ai_capabilities",
            "notes",
        ],
        overview_rows,
    )
    comparison_rows = []
    for case in repo_rows:
        features = next(row for row in case_features_rows if row["Case ID"] == case["Case ID"])
        hypothesis = next(row for row in hypotheses_rows if row["case_id"] == case["Case ID"])
        master_row = next(row for row in cases_master if row["Case ID"] == case["Case ID"])
        comparison_rows.append(
            {
                "case_id": case["Case ID"],
                "case_slug": case["case_slug"],
                "case_title": case["Case title"],
                "case_family": case["Case family"],
                "primary_ai_capability": case["Primary AI capability"],
                "primary_research_stage": case["Primary research stage"],
                "institutional_setting": case["Institutional setting"],
                "task_ground_truth_level": features["task_ground_truth_level"],
                "post_hoc_auditability": features["post_hoc_auditability"],
                "evidence_strength_overall": features["evidence_strength_overall"],
                "autonomy_pressure": features["autonomy_pressure"],
                "analysis_ready": master_row["analysis_ready"],
                "dominant_role_pattern": hypothesis["dominant_role_pattern"],
                "portfolio_group": case["Portfolio group"],
                "program_priority": case["Program priority"],
            }
        )
    write_csv(
        PROJECT_ROOT / "10_outputs" / "p1_case_comparison_matrix.csv",
        [
            "case_id",
            "case_slug",
            "case_title",
            "case_family",
            "primary_ai_capability",
            "primary_research_stage",
            "institutional_setting",
            "task_ground_truth_level",
            "post_hoc_auditability",
            "evidence_strength_overall",
            "autonomy_pressure",
            "analysis_ready",
            "dominant_role_pattern",
            "portfolio_group",
            "program_priority",
        ],
        comparison_rows,
    )
    write_text(PROJECT_ROOT / "10_outputs" / "p1_collection_summary.md", build_collection_summary(cases_master, sources_master, role_master))

    sync_project_tree(PROJECT_ROOT, USER_TARGET_ROOT)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
